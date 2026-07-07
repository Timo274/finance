from __future__ import annotations

import argparse
import math
from pathlib import Path
from typing import Any, Dict, List

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import yaml
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

LOWER_IS_BETTER = {"pe_ratio", "ev_ebitda", "debt_to_equity"}
HIGHER_IS_BETTER = {"roe", "revenue_growth", "fcf_growth", "gross_margin", "market_cap"}


def safe_float(value: Any) -> float | None:
    if value is None:
        return None
    try:
        value = float(value)
        if math.isnan(value) or math.isinf(value):
            return None
        return value
    except Exception:
        return None


def pct_change_from_series(series: pd.Series) -> float | None:
    vals = [safe_float(v) for v in series.dropna().tolist()]
    vals = [v for v in vals if v is not None]
    if len(vals) < 2:
        return None
    # yfinance columns are often newest first; compare newest vs previous
    newest, previous = vals[0], vals[1]
    if previous == 0:
        return None
    return newest / previous - 1


def fetch_company_metrics(ticker: str) -> Dict[str, Any]:
    """Fetch market and fundamental metrics for one public company."""
    asset = yf.Ticker(ticker)
    info = asset.info or {}

    market_cap = safe_float(info.get("marketCap"))
    enterprise_value = safe_float(info.get("enterpriseValue"))
    ebitda = safe_float(info.get("ebitda"))
    pe_ratio = safe_float(info.get("trailingPE"))
    roe = safe_float(info.get("returnOnEquity"))
    revenue_growth = safe_float(info.get("revenueGrowth"))
    gross_margin = safe_float(info.get("grossMargins"))
    total_debt = safe_float(info.get("totalDebt"))
    total_equity = None
    fcf_growth = None

    try:
        balance_sheet = asset.balance_sheet
        if "Stockholders Equity" in balance_sheet.index:
            total_equity = safe_float(balance_sheet.loc["Stockholders Equity"].dropna().iloc[0])
        elif "Common Stock Equity" in balance_sheet.index:
            total_equity = safe_float(balance_sheet.loc["Common Stock Equity"].dropna().iloc[0])
    except Exception:
        total_equity = None

    try:
        cash_flow = asset.cashflow
        if "Free Cash Flow" in cash_flow.index:
            fcf_growth = pct_change_from_series(cash_flow.loc["Free Cash Flow"])
    except Exception:
        fcf_growth = None

    ev_ebitda = enterprise_value / ebitda if enterprise_value and ebitda and ebitda > 0 else None
    debt_to_equity = total_debt / total_equity if total_debt is not None and total_equity and total_equity > 0 else None

    return {
        "ticker": ticker,
        "company": info.get("shortName") or info.get("longName") or ticker,
        "sector": info.get("sector"),
        "industry": info.get("industry"),
        "market_cap": market_cap,
        "pe_ratio": pe_ratio,
        "ev_ebitda": ev_ebitda,
        "roe": roe,
        "revenue_growth": revenue_growth,
        "fcf_growth": fcf_growth,
        "debt_to_equity": debt_to_equity,
        "gross_margin": gross_margin,
        "current_price": safe_float(info.get("currentPrice")),
        "recommendation": info.get("recommendationKey"),
    }


def passes_filters(row: pd.Series, filters: Dict[str, float]) -> bool:
    checks = [
        (row.get("market_cap"), ">=", filters.get("market_cap_min")),
        (row.get("pe_ratio"), "<=", filters.get("pe_max")),
        (row.get("ev_ebitda"), "<=", filters.get("ev_ebitda_max")),
        (row.get("roe"), ">=", filters.get("roe_min")),
        (row.get("revenue_growth"), ">=", filters.get("revenue_growth_min")),
        (row.get("fcf_growth"), ">=", filters.get("fcf_growth_min")),
        (row.get("debt_to_equity"), "<=", filters.get("debt_to_equity_max")),
        (row.get("gross_margin"), ">=", filters.get("gross_margin_min")),
    ]
    for value, op, threshold in checks:
        if threshold is None or pd.isna(threshold):
            continue
        if value is None or pd.isna(value):
            return False
        if op == ">=" and value < threshold:
            return False
        if op == "<=" and value > threshold:
            return False
    return True


def normalize_metric(series: pd.Series, higher_is_better: bool) -> pd.Series:
    series = pd.to_numeric(series, errors="coerce")
    if series.notna().sum() == 0:
        return pd.Series([0.5] * len(series), index=series.index)
    filled = series.fillna(series.median())
    low, high = filled.min(), filled.max()
    if high == low:
        return pd.Series([0.5] * len(series), index=series.index)
    score = (filled - low) / (high - low)
    return score if higher_is_better else 1 - score


def score_stocks(df: pd.DataFrame, weights: Dict[str, float]) -> pd.DataFrame:
    scored = df.copy()
    total_weight = sum(weights.values()) or 1.0
    score_components = []
    for metric, weight in weights.items():
        if metric not in scored.columns:
            continue
        higher = metric in HIGHER_IS_BETTER
        component_name = f"score_{metric}"
        scored[component_name] = normalize_metric(scored[metric], higher) * (weight / total_weight)
        score_components.append(component_name)
    scored["weighted_score"] = scored[score_components].sum(axis=1) * 100
    scored["rank"] = scored["weighted_score"].rank(method="dense", ascending=False).astype(int)
    return scored.sort_values(["passed_screen", "weighted_score"], ascending=[False, False])


def export_excel(df: pd.DataFrame, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        cols = [
            "rank", "ticker", "company", "sector", "industry", "passed_screen", "weighted_score",
            "market_cap", "pe_ratio", "ev_ebitda", "roe", "revenue_growth", "fcf_growth",
            "debt_to_equity", "gross_margin", "current_price", "recommendation"
        ]
        df[cols].to_excel(writer, sheet_name="Screen Results", index=False)
        df[df["passed_screen"]][cols].to_excel(writer, sheet_name="Passed Screen", index=False)
        df.sort_values("weighted_score", ascending=False).head(10)[cols].to_excel(writer, sheet_name="Top 10", index=False)
    wb = load_workbook(output_path)
    navy = "0B1F33"
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 28)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.column_letter in ["H"]:
                    cell.number_format = "$#,##0,,.0"
                elif cell.column_letter in ["I", "J", "N"]:
                    cell.number_format = "0.0x"
                elif cell.column_letter in ["K", "L", "M", "O"]:
                    cell.number_format = "0.0%"
                elif cell.column_letter == "G":
                    cell.number_format = "0.0"
                elif cell.column_letter == "P":
                    cell.number_format = "$0.00"
    wb.save(output_path)


def export_charts(df: pd.DataFrame, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    top = df.sort_values("weighted_score", ascending=False).head(15)
    fig_score = px.bar(
        top, x="weighted_score", y="ticker", orientation="h", color="passed_screen",
        title="Top Ranked Stocks by Weighted Score", text="weighted_score",
        hover_data=["company", "sector", "pe_ratio", "ev_ebitda", "roe", "revenue_growth"],
    )
    fig_score.update_layout(yaxis={"categoryorder": "total ascending"}, template="plotly_white")
    fig_scatter = px.scatter(
        df, x="ev_ebitda", y="revenue_growth", size="market_cap", color="weighted_score",
        hover_name="ticker", hover_data=["company", "pe_ratio", "roe", "gross_margin"],
        title="Growth vs. Valuation: EV/EBITDA vs Revenue Growth", template="plotly_white",
    )
    fig_table = go.Figure(data=[go.Table(
        header=dict(values=["Rank", "Ticker", "Company", "Score", "P/E", "EV/EBITDA", "ROE", "Rev Growth", "Passed"], fill_color="#0B1F33", font=dict(color="white"), align="left"),
        cells=dict(values=[
            df["rank"], df["ticker"], df["company"], df["weighted_score"].round(1),
            df["pe_ratio"].round(1), df["ev_ebitda"].round(1), (df["roe"]*100).round(1),
            (df["revenue_growth"]*100).round(1), df["passed_screen"]
        ], align="left")
    )])
    html = "<html><head><title>Stock Screener Charts</title></head><body>" + fig_score.to_html(full_html=False, include_plotlyjs="cdn") + fig_scatter.to_html(full_html=False, include_plotlyjs=False) + fig_table.to_html(full_html=False, include_plotlyjs=False) + "</body></html>"
    output_path.write_text(html, encoding="utf-8")


def run_screen(config_path: Path) -> pd.DataFrame:
    config = yaml.safe_load(config_path.read_text())
    tickers: List[str] = config["tickers"]
    rows = []
    for ticker in tickers:
        print(f"Fetching {ticker}...")
        try:
            rows.append(fetch_company_metrics(ticker))
        except Exception as exc:
            rows.append({"ticker": ticker, "company": ticker, "error": str(exc)})
    df = pd.DataFrame(rows)
    df["passed_screen"] = df.apply(lambda row: passes_filters(row, config.get("filters", {})), axis=1)
    df = score_stocks(df, config.get("weights", {}))
    output = config.get("output", {})
    base = config_path.parent.parent
    export_excel(df, base / output.get("excel_file", "screen_results.xlsx"))
    export_charts(df, base / output.get("chart_file", "screen_charts.html"))
    return df


def main() -> None:
    parser = argparse.ArgumentParser(description="Run a customizable investment stock screener.")
    parser.add_argument("--config", default="config/default_config.yaml", help="Path to YAML config file")
    args = parser.parse_args()
    config_path = Path(args.config).resolve()
    df = run_screen(config_path)
    print("\nTop ranked stocks:")
    print(df[["rank", "ticker", "company", "passed_screen", "weighted_score", "pe_ratio", "ev_ebitda", "roe", "revenue_growth"]].head(15).to_string(index=False))


if __name__ == "__main__":
    main()
