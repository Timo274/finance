from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

COLUMNS = [
    "rank",
    "ticker",
    "company",
    "sector",
    "industry",
    "passed_screen",
    "weighted_score",
    "market_cap",
    "pe_ratio",
    "ev_ebitda",
    "roe",
    "revenue_growth",
    "fcf_growth",
    "debt_to_equity",
    "gross_margin",
    "current_price",
    "recommendation",
    "error",
]


def export_excel(df: pd.DataFrame, path: str | Path) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    cols = [c for c in COLUMNS if c in df.columns]
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df[cols].to_excel(writer, sheet_name="Screen Results", index=False)
        df[df["passed_screen"]][cols].to_excel(writer, sheet_name="Passed Screen", index=False)
        df.sort_values("weighted_score", ascending=False).head(10)[cols].to_excel(
            writer, sheet_name="Top 10", index=False
        )
    wb = load_workbook(path)
    navy = "0B1F33"
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(
                max((len(str(c.value)) for c in col if c.value is not None), default=8) + 2, 30
            )
    wb.save(path)
